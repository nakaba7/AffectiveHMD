import numpy as np
from sklearn import datasets
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import confusion_matrix
import torch
import torch.nn.functional as F
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import DataLoader, TensorDataset
from torch.optim import SGD
import pandas as pd
import numpy as np
from EarlyStopping import EarlyStopping
import openpyxl
import itertools
import wandb
import statistics

SENSOR_NUM = 16
BATCH_SIZE = 64
OPTIMIZER_LEARNING_RATE = 1.0e-4
HIDDEN_DIM = 128
EPOCH_NUM = 500
SEQUENCE_TENSOR_LENGTH = 20
LABEL_SMOOTHING = 0.1
DROPOUT = 0.3
MODELPATH = "KFold_DNN_AllData.pth"
K = 10
book = openpyxl.Workbook()
xlsx_filename = "C:\\Users\\yukin\\Downloads\\MachineLearningResult\\AllData_DNN.xlsx"
book.save(xlsx_filename)
wb = openpyxl.load_workbook(xlsx_filename)
ws = wb['Sheet']

def macro_recall_score(cm):
    recall_list=[]
    for i in range(len(cm[0])):
        row_list=[]
        for j in range(len(cm[0])):
            row_list.append(cm[i][j])
        #print("sum", sum(row_list), "分母", cm[i][i])
        tmp_recall = cm[i][i] / sum(row_list) 
        #print("i",i,"recall", tmp_recall)
        recall_list.append(tmp_recall)
    return sum(recall_list)/len(recall_list)

def macro_precision_score(cm):
    precision_list=[]
    for i in range(len(cm[0])):
        column_list=[]
        for j in range(len(cm[0])):
            column_list.append(cm[j][i])
        #print("sum", sum(row_list), "分母", cm[i][i])
        tmp_precision = cm[i][i] / sum(column_list) 
        #print("i",i,"recall", tmp_recall)
        precision_list.append(tmp_precision)
    return sum(precision_list)/len(precision_list)

def macro_f1_score(cm):
    precision = macro_precision_score(cm)
    recall = macro_recall_score(cm)
    return 2 / (1/precision + 1/recall)

def zscore(x, axis = None):#標準化
    xmean = np.mean(x, axis=axis, keepdims=True)
    xstd  = np.std(x, axis=axis, keepdims=True)
    zscore = (x-xmean)/xstd

    return zscore

def mkDataSet(filename, headdatanum, is_normalize = True):#csvファイルデータをnumpyで返す
    """
    datasize : 訓練データのサイズ
    data_length : 過去何個分のデータを参考にするか
    train_x : トレーニングデータ（t=1,2,...,size-1の値)
    train_t : トレーニングデータのラベル（t=sizeの値）
    """
    print("dataset filename = {}".format(filename))
    df = pd.read_csv(filename, header=None)
    df = df.dropna(how="all", axis=0).dropna(how="all", axis=1)#Delete rows and columns with NaNs
    sensor_data = df.iloc[:, 1:SENSOR_NUM + headdatanum + 1]#get data without 1st row
    label_data = df.iloc[:, 0]#get label in 1st row

    value_data, label_data = sensor_data[:], label_data[:]

    data_size = label_data.to_numpy().shape[0]
    x_train_list = value_data.to_numpy().tolist()       
    y_train_list = label_data.to_numpy().tolist()
    
    print("datasize",data_size)
    train_x = []
    train_t = []
    
    for i in range(data_size):
       
        if y_train_list[i] == 'a':
            continue
        train_x.append(x_train_list[i])
        train_t.append(y_train_list[i])   
    train_x = np.array(train_x)
    #print(train_x.shape)
    train_t = np.array(train_t)
    train_t = train_t.astype(np.int32)
    #print(train_t[25958-data_length -3:25958-data_length +10 ])
    #print(train_x.shape)
    #print(train_x.shape[0], train_x.shape[1])
    if is_normalize == True:#列で正規化
        train_x = train_x.astype(np.float32)
        train_x = zscore(train_x, axis = 0)
        
    else:
      train_x = train_x.astype(np.float32)
        
    return train_x, train_t

def TrainandVal(filename, headdatanum, start_row, start_column):
    wandb.init(project="Affective HMD DNN")
    class NeuralNet(nn.Module):    
        def __init__(self):
            super(NeuralNet, self).__init__()
            self.input_layer = nn.Linear(SENSOR_NUM+headdatanum, HIDDEN_DIM)
            self.hidden_layer1 = nn.Linear(HIDDEN_DIM, HIDDEN_DIM)
            self.hidden_layer2 = nn.Linear(HIDDEN_DIM, 64)
            self.output_layer = nn.Linear(64, 5)
            self.dropout = nn.Dropout(DROPOUT)
            self.bn1 = nn.BatchNorm1d(HIDDEN_DIM)
            self.bn2 = nn.BatchNorm1d(64)
        
        def forward(self, x):
            x = self.input_layer(x)
            x = F.relu(x)
            x = self.dropout(x)
            x = self.hidden_layer1(x)
            x = self.bn1(x)#バッチ正規化
            x = F.relu(x)
            #x = self.dropout(x)
            x = self.hidden_layer2(x)
            x = self.bn2(x)
            x = F.relu(x)
            #x = self.dropout(x)
            output = self.output_layer(x)
            #x = F.softmax(x, dim=1)
            return output
    data, target = mkDataSet(filename, headdatanum, is_normalize=False)
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    train_acc_list = []
    val_acc_list = []
    test_acc_list = []
    prec_list = []
    recall_list = []
    f_list = []

    for i in range(K):#10回で交差検証
        train_val_sensor, test_sensor, train_val_label, test_label = train_test_split(data, target, random_state=i, test_size=0.2)#訓練＋評価とテストに分ける
        val_size = int(test_label.shape[0]/2)
        train_sensor, val_sensor, train_label, val_label = train_test_split(train_val_sensor, train_val_label, random_state=0, test_size=val_size)#訓練と評価に分ける

        print("訓練データ", train_sensor.shape)
        print("評価データ", val_sensor.shape)
        print("テストデータ", test_sensor.shape)

        train_sensor = torch.from_numpy(train_sensor).float()
        train_label = torch.from_numpy(train_label).long()
        val_sensor = torch.from_numpy(val_sensor).float()
        val_label = torch.from_numpy(val_label).long()
        test_sensor = torch.from_numpy(test_sensor).float()
        test_label = torch.from_numpy(test_label).long()

        train_label = torch.nn.functional.one_hot(train_label, num_classes=5)
        val_label = torch.nn.functional.one_hot(val_label, num_classes=5)
        test_label = torch.nn.functional.one_hot(test_label, num_classes=5)

        # 2. Datasetの作成
        train_dataset = TensorDataset(train_sensor, train_label)
        valid_dataset = TensorDataset(val_sensor, val_label)
        test_dataset = TensorDataset(test_sensor, test_label)

        # 3. DataLoaderの作成
        batch_size = BATCH_SIZE
        train_dataloader = DataLoader(train_dataset, batch_size=batch_size, shuffle=True)
        valid_dataloader = DataLoader(valid_dataset, batch_size=batch_size, shuffle=True)
        test_dataloader = DataLoader(test_dataset, batch_size=batch_size, shuffle=True)

        model = NeuralNet().to(device)
        # 5. 損失関数の定義
        criterion = nn.CrossEntropyLoss(label_smoothing=LABEL_SMOOTHING)

        # 6. 最適化手法の定義
        optimizer = optim.Adam(model.parameters(), lr=OPTIMIZER_LEARNING_RATE)
        # 7. 学習・評価
        # エポック数
        num_epochs = EPOCH_NUM

        # 学習時と検証時で分けるためディクショナリを用意
        dataloaders_dict = {
            'train': train_dataloader,
            'val': valid_dataloader
        }

        earlystopping = EarlyStopping(patience=30, verbose=True, path=MODELPATH)
        # 損失和
        epoch_loss = 0.0
        val_epoch_loss = 0.0
        # 正解数
        epoch_corrects = 0
        val_epoch_corrects = 0

        last_train_acc = 0
        last_val_acc = 0
        min_val_loss = 500
        last_epoch_num = 0
        #last_pred_list = []
        #last_labels_list = []
        for epoch in range(num_epochs):
            print('Epoch {}/{}'.format(epoch+1, num_epochs))
            print('-------------')
            #pred_list = []
            #labels_list = []
            for phase in ['train', 'val']:
                
                if phase == 'train':
                    # モデルを訓練モードに設定
                    model.train()
                else:
                    # モデルを推論モードに設定
                    model.eval()
                
                # 損失和
                epoch_loss = 0.0
                val_epoch_loss = 0.0

                # 正解数
                epoch_corrects = 0
                val_epoch_corrects = 0
                
                # DataLoaderからデータをバッチごとに取り出す
                for inputs, labels in dataloaders_dict[phase]:
                    
                    # optimizerの初期化
                    optimizer.zero_grad()
                    
                    # 学習時のみ勾配を計算させる設定にする
                    with torch.set_grad_enabled(phase == 'train'):
                        outputs = model(inputs.to(device))
                        labels = labels.to(torch.float32)
                        # 損失を計算
                        loss = criterion(outputs.to(device), labels.to(device))
                        
                        # ラベルを予測
                        _, preds = torch.max(outputs, 1)
                        _, answer = torch.max(labels,1)
                        # 訓練時はバックプロパゲーション
                        if phase == 'train':
                            # 逆伝搬の計算
                            loss.backward()
                            # パラメータの更新
                            optimizer.step()
                        
                        # イテレーション結果の計算
                        # lossの合計を更新
                        # PyTorchの仕様上各バッチ内での平均のlossが計算される。
                        # データ数を掛けることで平均から合計に変換をしている。
                        # 損失和は「全データの損失/データ数」で計算されるため、
                        # 平均のままだと損失和を求めることができないため。
                        if phase == 'train':
                            epoch_loss += loss.item() * inputs.size(0)
                        else:
                            val_epoch_loss += loss.item() * inputs.size(0)

                        preds = preds.to(device)
                        answer = answer.to(device)
                        # 正解数の合計を更新
                        if phase == 'train':
                            epoch_corrects += torch.sum(preds == answer)
                        else:
                            val_epoch_corrects += torch.sum(preds == answer)
                            #pred_list.append(preds.tolist())
                            #labels_list.append(answer.tolist())

                # epochごとのlossと正解率を表示
                if phase == 'train':
                    epoch_loss = epoch_loss / len(dataloaders_dict[phase].dataset)
                    epoch_acc = epoch_corrects.double() / len(dataloaders_dict[phase].dataset)
                    last_train_acc = epoch_acc
                    wandb.log({'epoch': epoch, 'train_loss': epoch_loss, 'train_Accuracy': epoch_acc})
                    print('{} Loss: {:.4f} Acc: {:.4f}'.format(phase, epoch_loss, epoch_acc))
                else:
                    val_epoch_loss = val_epoch_loss / len(dataloaders_dict[phase].dataset)
                    val_epoch_acc = val_epoch_corrects.double() / len(dataloaders_dict[phase].dataset)
                    
                    wandb.log({'epoch': epoch, 'val_loss': val_epoch_loss, 'val_Accuracy': val_epoch_acc})
                    print('{} Loss: {:.4f} Acc: {:.4f}'.format(phase, val_epoch_loss, val_epoch_acc))
                    if val_epoch_loss < min_val_loss:
                        #print("-------------best loss : ",val_epoch_loss)
                        #print("-------------best acc : ",val_epoch_acc.item())
                        last_train_acc = epoch_acc
                        last_val_acc = val_epoch_acc
                        min_val_loss = val_epoch_loss
                        last_epoch_num = epoch
                        #last_pred_list = pred_list.copy()
                        #last_labels_list = labels_list.copy()
            earlystopping(val_epoch_loss, model) #callメソッド呼び出し
            if earlystopping.early_stop: #ストップフラグがTrueの場合、breakでforループを抜ける
                print("Early Stopping!")
                print("val acc : ",last_val_acc.item())
                break
            if epoch == num_epochs-1:
                last_train_acc = epoch_acc
                last_val_acc = val_epoch_acc
                min_val_loss = val_epoch_loss
                last_epoch_num = epoch
        #------------------------------------------------------------テスト開始----------------------------------------------------------------
        model.load_state_dict(torch.load(MODELPATH))
        model.eval()
        pred_list = []
        labels_list = []
        test_corrects_num = 0
        for inputs, labels in test_dataloader:
            outputs = model(inputs.to(device))
            labels = labels.to(torch.float32)
            loss = criterion(outputs.to(device), labels.to(device))
            # ラベルを予測
            labels.to(device)
            _, preds = torch.max(outputs, 1)
            _, answer = torch.max(labels,1)
            preds = preds.to(device)
            answer = answer.to(device)
            test_corrects_num += torch.sum(preds==answer)#正解数
            pred_list.append(preds.tolist())
            labels_list.append(answer.tolist())
        test_acc = test_corrects_num.double() / len(test_dataloader.dataset)
        cm = confusion_matrix(list(itertools.chain.from_iterable(pred_list)), list(itertools.chain.from_iterable(labels_list)))
        #print("confusion_matrix", cm)

        prec = macro_precision_score(cm)
        recall = macro_recall_score(cm)
        f = macro_f1_score(cm)
        last_train_acc = last_train_acc.item()
        last_val_acc = last_val_acc.item()
        test_acc = test_acc.item()

        train_acc_list.append(last_train_acc)
        val_acc_list.append(last_val_acc)
        test_acc_list.append(test_acc)
        prec_list.append(prec)
        recall_list.append(recall)
        f_list.append(f)

        wandb.save("NN_DNN.h5")
        print()
        print("DNN")
        if headdatanum == 2:
            print("Head Rotation Data Included")
        else:
            print("No Head Rotation Data")
        print("------------------------------------------")
        print("AllDNN",i+1,"回目")
        print("訓練データ精度：", last_train_acc)
        print("評価データ精度：", last_val_acc)
        print("テストデータ精度：", test_acc)
        print("適合率", prec)        
        print("再現率", recall)
        print("F1", f)
        print("------------------------------------------")
        print("バッチサイズ：",BATCH_SIZE)
        print("学習率：",OPTIMIZER_LEARNING_RATE)
        print("エポック数：",last_epoch_num)
        print("Confusion Matrix",cm)

        #配列宣言
        eval = ['DNN','訓練データ精度','評価データ精度','テストデータ', '適合率','再現率','F1']
        Headdata = ['頭部姿勢あり', last_train_acc, last_val_acc, test_acc, prec, recall, f]
        NoHeaddata = ['頭部姿勢なし',last_train_acc, last_val_acc, test_acc, prec, recall, f]
        
        #配列ループ
        ws.cell(start_row, start_column, "AllDNN")
        ws.cell(start_row, start_column+1, i)
        for j in range(0,len(eval)):
            #A列にリストを書き込み
            ws.cell(start_row+j+1,start_column,value = eval[j])
            #B列にリストを書き込み
            if headdatanum == 2:
                ws.cell(start_row+j+1,start_column+1,value = Headdata[j])
            else:
                ws.cell(start_row+j+1,start_column+2, value = NoHeaddata[j])
        ws.cell(start_row+8, start_column, value='エポック数')
        ws.cell(start_row+8, start_column+1, value=last_epoch_num+1)
        wb.save(xlsx_filename)
        start_row += 10
        if i == K-1:#K-分割の最終結果書き込み
            train_mean = statistics.mean(train_acc_list)
            val_mean = statistics.mean(val_acc_list)
            test_mean = statistics.mean(test_acc_list)
            prec_mean = statistics.mean(prec_list)
            recall_mean = statistics.mean(recall_list)
            f_mean = statistics.mean(f_list)
            meanHeaddata = ['頭部姿勢あり', train_mean, val_mean, test_mean, prec_mean, recall_mean, f_mean]
            meanNoHeaddata = ['頭部姿勢なし', train_mean, val_mean, test_mean, prec_mean, recall_mean, f_mean]
            start_row+=2
            ws.cell(start_row, start_column, "AllDNN")
            ws.cell(start_row, start_column+1, '平均')
            for i in range(0,len(eval)):
                #A列にリストを書き込み
                ws.cell(start_row+i+1,start_column,value = eval[i])

                #B列にリストを書き込み
                if headdatanum == 2:
                    ws.cell(start_row+i+1,start_column+1,value = meanHeaddata[i])
                else:
                    ws.cell(start_row+i+1,start_column+2, value = meanNoHeaddata[i])
            


column_count=1
row_count=1
filename="C:\\Users\\yukin\\Downloads\\All10Dataset.csv"
TrainandVal(filename, 2, row_count, column_count)
TrainandVal(filename, 0, row_count, column_count)
column_count += 4






