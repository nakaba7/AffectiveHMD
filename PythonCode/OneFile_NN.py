import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.metrics import confusion_matrix
import torch
import torch.nn.functional as F
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import DataLoader, TensorDataset
from torch.optim import SGD
import pandas as pd
import numpy as np

from EarlyStopping import EarlyStopping
from tqdm import tqdm
import itertools
import openpyxl

SENSOR_NUM = 16
BATCH_SIZE = 64
LEARNING_RATE = 1.0e-4
HIDDEN_DIM = 64
EPOCH_NUM = 500
DROPOUT=0.3

book = openpyxl.Workbook()
xlsx_filename = "C:\\Users\\yukin\\Downloads\\MachineLearningResult\\AllData_NN.xlsx"
book.save(xlsx_filename)
wb = openpyxl.load_workbook(xlsx_filename)
ws = wb['Sheet']
import wandb


def macro_recall_score(cm):
    recall_list=[]
    for i in range(len(cm[0])):
        row_list=[]
        for j in range(len(cm[0])):
            row_list.append(cm[i][j])
        #print("sum", sum(row_list), "分母", cm[i][i])
        tmp_recall = cm[i][i] / sum(row_list) 
        #print("i",i,"recall", tmp_recall)
        recall_list.append(tmp_recall)
    return sum(recall_list)/len(recall_list)

def macro_precision_score(cm):
    precision_list=[]
    for i in range(len(cm[0])):
        column_list=[]
        for j in range(len(cm[0])):
            column_list.append(cm[j][i])
        #print("sum", sum(row_list), "分母", cm[i][i])
        tmp_precision = cm[i][i] / sum(column_list) 
        #print("i",i,"recall", tmp_recall)
        precision_list.append(tmp_precision)
    return sum(precision_list)/len(precision_list)

def macro_f1_score(cm):
    precision = macro_precision_score(cm)
    recall = macro_recall_score(cm)
    return 2 / (1/precision + 1/recall)

def zscore(x, axis = None):#標準化
    xmean = np.mean(x, axis=axis, keepdims=True)
    xstd  = np.std(x, axis=axis, keepdims=True)
    zscore = (x-xmean)/xstd
    print("mean",np.squeeze(xmean))
    print("std",np.squeeze(xstd))
    return zscore

def mkDataSet(filename, headdatanum, is_normalize = True):#csvファイルデータをnumpyで返す
    """
    datasize : 訓練データのサイズ
    data_length : 過去何個分のデータを参考にするか
    train_x : トレーニングデータ（t=1,2,...,size-1の値)
    train_t : トレーニングデータのラベル（t=sizeの値）
    """
    print("dataset filename = {}".format(filename))
    df = pd.read_csv(filename, header=None)
    df = df.dropna(how="all", axis=0).dropna(how="all", axis=1)#Delete rows and columns with NaNs
    sensor_data = df.iloc[:, 1:SENSOR_NUM + headdatanum + 1]#get data without 1st row
    label_data = df.iloc[:, 0]#get label in 1st row

    value_data, label_data = sensor_data[:], label_data[:]

    data_size = label_data.to_numpy().shape[0]
    x_train_list = value_data.to_numpy().tolist()       
    y_train_list = label_data.to_numpy().tolist()
    
    print("datasize",data_size)
    train_x = []
    train_t = []
    
    for i in range(data_size):
       
        if y_train_list[i] == 'a':
            continue
        train_x.append(x_train_list[i])
        train_t.append(y_train_list[i])   
    train_x = np.array(train_x)
    #print(train_x.shape)
    train_t = np.array(train_t)
    train_t = train_t.astype(np.int32)
    #print(train_t[25958-data_length -3:25958-data_length +10 ])
    #print(train_x.shape)
    #print(train_x.shape[0], train_x.shape[1])
    if is_normalize == True:#列で正規化
        train_x = train_x.astype(np.float32)
        train_x = zscore(train_x, axis = 0)
        
    else:
      train_x = train_x.astype(np.float32)
        
    return train_x, train_t

def TrainandVal(train_filename, headdatanum, start_row, start_column):
    wandb.init(project="Affective HMD NN")
    class NeuralNet(nn.Module):    
        def __init__(self):
            super(NeuralNet, self).__init__()
            self.input_layer = nn.Linear(SENSOR_NUM+headdatanum, HIDDEN_DIM)
            self.hidden_layer1 = nn.Linear(HIDDEN_DIM, HIDDEN_DIM)
            self.hidden_layer2 = nn.Linear(HIDDEN_DIM, 64)
            self.output_layer = nn.Linear(64, 5)
            self.dropout = nn.Dropout(DROPOUT)
            self.bn1 = nn.BatchNorm1d(HIDDEN_DIM)
            self.bn2 = nn.BatchNorm1d(64)
        
        def forward(self, x):
            x = self.input_layer(x)
            x = F.relu(x)
            x = self.dropout(x)
            x = self.hidden_layer1(x)
            x = self.bn1(x)#バッチ正規化
            x = F.relu(x)
            #x = self.dropout(x)
            x = self.hidden_layer2(x)
            x = self.bn2(x)
            x = F.relu(x)
            #x = self.dropout(x)
            output = self.output_layer(x)
            #x = F.softmax(x, dim=1)
            return output
    data, target = mkDataSet(train_filename, headdatanum, is_normalize=True)
    x_train = data
    y_train = target
    #x_valid, y_valid = mkDataSet(val_filename, headdatanum, is_normalize=True)

    # 学習データと検証データに分割
    x_train, x_valid, y_train, y_valid = train_test_split(data, target, shuffle=True, random_state=0)

    print("x_train",x_train.shape)
    print("y_train",y_train.shape)
    print("x_valid",x_valid.shape)
    print("y_valid",y_valid.shape)

    # Tensor型に変換
    # 学習に入れるときはfloat型になっている必要があるのここで変換してしまう
    x_train = torch.from_numpy(x_train).float()
    y_train = torch.from_numpy(y_train).long()
    x_valid = torch.from_numpy(x_valid).float()
    y_valid = torch.from_numpy(y_valid).long()

    #print('x_train : ', x_train.shape)
    #print('y_train : ', y_train.shape)
    #print('x_valid : ', x_valid.shape)
    #print('y_valid : ', y_valid.shape)

    y_train = torch.nn.functional.one_hot(y_train, num_classes=5)
    y_valid = torch.nn.functional.one_hot(y_valid, num_classes=5)

    # 2. Datasetの作成
    train_dataset = TensorDataset(x_train, y_train)
    valid_dataset = TensorDataset(x_valid, y_valid)

    # 3. DataLoaderの作成
    batch_size = BATCH_SIZE
    train_dataloader = DataLoader(train_dataset, batch_size=batch_size, shuffle=True)
    valid_dataloader = DataLoader(valid_dataset, batch_size=batch_size, shuffle=True)

    #model = torch.nn.DataParallel(model)
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    model = NeuralNet().to(device)
    model = torch.nn.DataParallel(model)
    # 5. 損失関数の定義
    criterion = nn.CrossEntropyLoss()

    # 6. 最適化手法の定義
    #optimizer = optim.SGD(model.parameters(), lr=LEARNING_RATE)
    optimizer = optim.Adam(model.parameters(), lr=LEARNING_RATE)
    # 7. 学習・評価
    # エポック数
    num_epochs = EPOCH_NUM

    # 学習時と検証時で分けるためディクショナリを用意
    dataloaders_dict = {
        'train': train_dataloader,
        'val': valid_dataloader
    }


    earlystopping = EarlyStopping(patience=30, verbose=True, path="NN_Data.pth")
    # 損失和
    epoch_loss = 0.0
    val_epoch_loss = 0.0
    # 正解数
    epoch_corrects = 0
    val_epoch_corrects = 0

    last_train_acc = 0
    last_test_acc = 0
    min_val_loss = 500
    last_pred_list = []
    last_labels_list = []
    for epoch in range(num_epochs):
        print('Epoch {}/{}'.format(epoch+1, num_epochs))
        print('-------------')
        pred_list = []
        labels_list = []
        for phase in ['train', 'val']:
            
            if phase == 'train':
                # モデルを訓練モードに設定
                model.train()
            else:
                # モデルを推論モードに設定
                model.eval()
            
            # 損失和
            epoch_loss = 0.0
            val_epoch_loss = 0.0

            # 正解数
            epoch_corrects = 0
            val_epoch_corrects = 0
            
            # DataLoaderからデータをバッチごとに取り出す
            for inputs, labels in dataloaders_dict[phase]:
                
                # optimizerの初期化
                optimizer.zero_grad()
                
                # 学習時のみ勾配を計算させる設定にする
                with torch.set_grad_enabled(phase == 'train'):
                    outputs = model(inputs.to(device))
                    labels = labels.to(torch.float32)
                    # 損失を計算
                    loss = criterion(outputs.to(device), labels.to(device))
                    
                    # ラベルを予測
                    _, preds = torch.max(outputs, 1)
                    _, answer = torch.max(labels,1)
                    # 訓練時はバックプロパゲーション
                    if phase == 'train':
                        # 逆伝搬の計算
                        loss.backward()
                        # パラメータの更新
                        optimizer.step()
                    
                    # イテレーション結果の計算
                    # lossの合計を更新
                    # PyTorchの仕様上各バッチ内での平均のlossが計算される。
                    # データ数を掛けることで平均から合計に変換をしている。
                    # 損失和は「全データの損失/データ数」で計算されるため、
                    # 平均のままだと損失和を求めることができないため。
                    if phase == 'train':
                        epoch_loss += loss.item() * inputs.size(0)
                    else:
                        val_epoch_loss += loss.item() * inputs.size(0)

                    preds = preds.to(device)
                    answer = answer.to(device)
                    # 正解数の合計を更新
                    if phase == 'train':
                        epoch_corrects += torch.sum(preds == answer)
                    else:
                        val_epoch_corrects += torch.sum(preds == answer)
                        pred_list.append(preds.tolist())
                        labels_list.append(answer.tolist())

            # epochごとのlossと正解率を表示
            if phase == 'train':
                epoch_loss = epoch_loss / len(dataloaders_dict[phase].dataset)
                epoch_acc = epoch_corrects.double() / len(dataloaders_dict[phase].dataset)
                last_train_acc = epoch_acc
                #wandb.log({'epoch': epoch, 'train_loss': epoch_loss, 'train_Accuracy': epoch_acc})
                print('{} Loss: {:.4f} Acc: {:.4f}'.format(phase, epoch_loss, epoch_acc))
            else:
                val_epoch_loss = val_epoch_loss / len(dataloaders_dict[phase].dataset)
                val_epoch_acc = val_epoch_corrects.double() / len(dataloaders_dict[phase].dataset)
                
                #wandb.log({'epoch': epoch, 'val_loss': val_epoch_loss, 'val_Accuracy': val_epoch_acc})
                print('{} Loss: {:.4f} Acc: {:.4f}'.format(phase, val_epoch_loss, val_epoch_acc))
                if val_epoch_loss < min_val_loss:
                    print("-------------best loss : ",val_epoch_loss)
                    print("-------------best acc : ",val_epoch_acc.item())
                    last_test_acc = val_epoch_acc
                    min_val_loss = val_epoch_loss
                    last_pred_list = pred_list.copy()
                    last_labels_list = labels_list.copy()
        earlystopping(val_epoch_loss, model) #callメソッド呼び出し
        if earlystopping.early_stop: #ストップフラグがTrueの場合、breakでforループを抜ける
            print("Early Stopping!")
            break


    cm = confusion_matrix(list(itertools.chain.from_iterable(last_pred_list)), list(itertools.chain.from_iterable(last_labels_list)))

    print("confusion_matrix", cm)

    last_train_acc = last_train_acc.item()
    last_test_acc = last_test_acc.item()

    prec = macro_precision_score(cm)
    recall = macro_recall_score(cm)
    f = macro_f1_score(cm)
    #wandb.save("BasicMachineLearning.h5")
    print()
    print("NN")
    if headdatanum == 2:
        print("Head Rotation Data Included")
    else:
        print("No Head Rotation Data")
    print("------------------------------------------")
    print("訓練データ精度：", last_train_acc)
    print("評価データ精度：", last_test_acc)
    print("適合率", prec)        
    print("再現率", recall)
    print("F1", f)
    print("------------------------------------------")
    print("バッチサイズ：",BATCH_SIZE)
    print("学習率：",LEARNING_RATE)
    print("エポック数：",EPOCH_NUM)
    print("Confusion Matrix",cm)

    #配列宣言
    eval = ['NN','訓練データ精度','評価データ精度','適合率','再現率','F1']
    Headdata = ['頭部姿勢あり', last_train_acc, last_test_acc, prec, recall, f]
    NoHeaddata = ['頭部姿勢なし',last_train_acc, last_test_acc, prec, recall, f]
    
    #配列ループ
    #ws.cell(start_row, start_column, name)
    for i in range(0,len(eval)):

        #A列にリストを書き込み
        ws.cell(start_row+i+1,start_column,value = eval[i])

        #B列にリストを書き込み
        if headdatanum == 2:
            ws.cell(start_row+i+1,start_column+1,value = Headdata[i])
        else:
            ws.cell(start_row+i+1,start_column+2, value = NoHeaddata[i])

    wb.save(xlsx_filename)



column_count=1
train_filename="C:\\Users\\yukin\\Downloads\\Relabeled_Nakabayashi_DataSet.csv"
#val_filename = "C:\\Users\\yukin\\Downloads\\Relabeled_Ozaki_DataSet.csv"
row_count=1
#TrainandVal(train_filename, 2, row_count, column_count)
TrainandVal(train_filename, 0, row_count, column_count)
    







