from sklearn.model_selection import train_test_split
from sklearn.metrics import confusion_matrix
import torch
import torch.nn as nn
import torch.optim as optim
from torch.utils.data import DataLoader, TensorDataset
from EarlyStopping import EarlyStopping
import openpyxl
import itertools
import wandb
import statistics
from Evaluation_Metric import macro_precision_score, macro_recall_score, macro_f1_score
from MakeDataSet import mkDataSet
from RNNModel import RNNClassifier

"""
時系列学習モデルのRNNで表情識別を行う.
K分割交差検証を行う.
"""

BATCH_SIZE = 64
OPTIMIZER_LEARNING_RATE = 1.0e-4
EPOCH_NUM = 500
LABEL_SMOOTHING = 0.1
MODELPATH = "KFold_RNN_AllData.pth"
K = 10
inputfile_name=".\\DataSet\\All10Dataset.csv" #学習するデータセットのcsvファイル
book = openpyxl.Workbook()
xlsx_filename = ".\\MachineLearningResult\\AllData_RNN.xlsx" #学習結果を書き込むExcelファイル
book.save(xlsx_filename)
wb = openpyxl.load_workbook(xlsx_filename)
ws = wb['Sheet']


def TrainandVal(inputfile_name, headdatanum, start_row, start_column):
    wandb.init(project="Affective HMD RNN")
    
    data, target = mkDataSet(inputfile_name, headdatanum, is_normalize=False)
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    train_acc_list = []
    val_acc_list = []
    test_acc_list = []
    prec_list = []
    recall_list = []
    f_list = []

    for i in range(K):#10回で交差検証
        train_val_sensor, test_sensor, train_val_label, test_label = train_test_split(data, target, random_state=i, test_size=0.2)#訓練＋評価とテストに分ける
        val_size = int(test_label.shape[0]/2)
        train_sensor, val_sensor, train_label, val_label = train_test_split(train_val_sensor, train_val_label, random_state=0, test_size=val_size)#訓練と評価に分ける

        print("訓練データ", train_sensor.shape)
        print("評価データ", val_sensor.shape)
        print("テストデータ", test_sensor.shape)

        train_sensor = torch.from_numpy(train_sensor).float()
        train_label = torch.from_numpy(train_label).long()
        val_sensor = torch.from_numpy(val_sensor).float()
        val_label = torch.from_numpy(val_label).long()
        test_sensor = torch.from_numpy(test_sensor).float()
        test_label = torch.from_numpy(test_label).long()

        train_label = torch.nn.functional.one_hot(train_label, num_classes=5)
        val_label = torch.nn.functional.one_hot(val_label, num_classes=5)
        test_label = torch.nn.functional.one_hot(test_label, num_classes=5)

        # 2. Datasetの作成
        train_dataset = TensorDataset(train_sensor, train_label)
        valid_dataset = TensorDataset(val_sensor, val_label)
        test_dataset = TensorDataset(test_sensor, test_label)

        # 3. DataLoaderの作成
        batch_size = BATCH_SIZE
        train_dataloader = DataLoader(train_dataset, batch_size=batch_size, shuffle=True)
        valid_dataloader = DataLoader(valid_dataset, batch_size=batch_size, shuffle=True)
        test_dataloader = DataLoader(test_dataset, batch_size=batch_size, shuffle=True)

        model = RNNClassifier().to(device)
        # 5. 損失関数の定義
        criterion = nn.CrossEntropyLoss(label_smoothing=LABEL_SMOOTHING)

        # 6. 最適化手法の定義
        optimizer = optim.Adam(model.parameters(), lr=OPTIMIZER_LEARNING_RATE)
        # 7. 学習・評価
        # エポック数
        num_epochs = EPOCH_NUM

        # 学習時と検証時で分けるためディクショナリを用意
        dataloaders_dict = {
            'train': train_dataloader,
            'val': valid_dataloader
        }

        earlystopping = EarlyStopping(patience=30, verbose=True, path=MODELPATH)
        # 損失和
        epoch_loss = 0.0
        val_epoch_loss = 0.0
        # 正解数
        epoch_corrects = 0
        val_epoch_corrects = 0

        last_train_acc = 0
        last_val_acc = 0
        min_val_loss = 500
        last_epoch_num = 0
        #last_pred_list = []
        #last_labels_list = []
        for epoch in range(num_epochs):
            print('Epoch {}/{}'.format(epoch+1, num_epochs))
            print('-------------')
            #pred_list = []
            #labels_list = []
            for phase in ['train', 'val']:
                
                if phase == 'train':
                    # モデルを訓練モードに設定
                    model.train()
                else:
                    # モデルを推論モードに設定
                    model.eval()
                
                # 損失和
                epoch_loss = 0.0
                val_epoch_loss = 0.0

                # 正解数
                epoch_corrects = 0
                val_epoch_corrects = 0
                
                # DataLoaderからデータをバッチごとに取り出す
                for inputs, labels in dataloaders_dict[phase]:
                    
                    # optimizerの初期化
                    optimizer.zero_grad()
                    
                    # 学習時のみ勾配を計算させる設定にする
                    with torch.set_grad_enabled(phase == 'train'):
                        outputs = model(inputs.to(device))
                        labels = labels.to(torch.float32)
                        # 損失を計算
                        loss = criterion(outputs.to(device), labels.to(device))
                        
                        # ラベルを予測
                        _, preds = torch.max(outputs, 1)
                        _, answer = torch.max(labels,1)
                        # 訓練時はバックプロパゲーション
                        if phase == 'train':
                            # 逆伝搬の計算
                            loss.backward()
                            # パラメータの更新
                            optimizer.step()
                        
                        # イテレーション結果の計算
                        # lossの合計を更新
                        # PyTorchの仕様上各バッチ内での平均のlossが計算される。
                        # データ数を掛けることで平均から合計に変換をしている。
                        # 損失和は「全データの損失/データ数」で計算されるため、
                        # 平均のままだと損失和を求めることができないため。
                        if phase == 'train':
                            epoch_loss += loss.item() * inputs.size(0)
                        else:
                            val_epoch_loss += loss.item() * inputs.size(0)

                        preds = preds.to(device)
                        answer = answer.to(device)
                        # 正解数の合計を更新
                        if phase == 'train':
                            epoch_corrects += torch.sum(preds == answer)
                        else:
                            val_epoch_corrects += torch.sum(preds == answer)
                            #pred_list.append(preds.tolist())
                            #labels_list.append(answer.tolist())

                # epochごとのlossと正解率を表示
                if phase == 'train':
                    epoch_loss = epoch_loss / len(dataloaders_dict[phase].dataset)
                    epoch_acc = epoch_corrects.double() / len(dataloaders_dict[phase].dataset)
                    last_train_acc = epoch_acc
                    wandb.log({'epoch': epoch, 'train_loss': epoch_loss, 'train_Accuracy': epoch_acc})
                    print('{} Loss: {:.4f} Acc: {:.4f}'.format(phase, epoch_loss, epoch_acc))
                else:
                    val_epoch_loss = val_epoch_loss / len(dataloaders_dict[phase].dataset)
                    val_epoch_acc = val_epoch_corrects.double() / len(dataloaders_dict[phase].dataset)
                    
                    wandb.log({'epoch': epoch, 'val_loss': val_epoch_loss, 'val_Accuracy': val_epoch_acc})
                    print('{} Loss: {:.4f} Acc: {:.4f}'.format(phase, val_epoch_loss, val_epoch_acc))
                    if val_epoch_loss < min_val_loss:
                        #print("-------------best loss : ",val_epoch_loss)
                        #print("-------------best acc : ",val_epoch_acc.item())
                        last_train_acc = epoch_acc
                        last_val_acc = val_epoch_acc
                        min_val_loss = val_epoch_loss
                        last_epoch_num = epoch
                        #last_pred_list = pred_list.copy()
                        #last_labels_list = labels_list.copy()
            earlystopping(val_epoch_loss, model) #callメソッド呼び出し
            if earlystopping.early_stop: #ストップフラグがTrueの場合、breakでforループを抜ける
                print("Early Stopping!")
                print("val acc : ",last_val_acc.item())
                break
            if epoch == num_epochs-1:
                last_train_acc = epoch_acc
                last_val_acc = val_epoch_acc
                min_val_loss = val_epoch_loss
                last_epoch_num = epoch
        #------------------------------------------------------------テスト開始----------------------------------------------------------------
        model.load_state_dict(torch.load(MODELPATH))
        model.eval()
        pred_list = []
        labels_list = []
        test_corrects_num = 0
        for inputs, labels in test_dataloader:
            outputs = model(inputs.to(device))
            labels = labels.to(torch.float32)
            loss = criterion(outputs.to(device), labels.to(device))
            # ラベルを予測
            labels.to(device)
            _, preds = torch.max(outputs, 1)
            _, answer = torch.max(labels,1)
            preds = preds.to(device)
            answer = answer.to(device)
            test_corrects_num += torch.sum(preds==answer)#正解数
            pred_list.append(preds.tolist())
            labels_list.append(answer.tolist())
        test_acc = test_corrects_num.double() / len(test_dataloader.dataset)
        cm = confusion_matrix(list(itertools.chain.from_iterable(pred_list)), list(itertools.chain.from_iterable(labels_list)))
        #print("confusion_matrix", cm)

        prec = macro_precision_score(cm)
        recall = macro_recall_score(cm)
        f = macro_f1_score(cm)
        last_train_acc = last_train_acc.item()
        last_val_acc = last_val_acc.item()
        test_acc = test_acc.item()

        train_acc_list.append(last_train_acc)
        val_acc_list.append(last_val_acc)
        test_acc_list.append(test_acc)
        prec_list.append(prec)
        recall_list.append(recall)
        f_list.append(f)

        wandb.save("NN_RNN.h5")
        print()
        print("RNN")
        if headdatanum == 2:
            print("Head Rotation Data Included")
        else:
            print("No Head Rotation Data")
        print("------------------------------------------")
        print("All",i+1,"回目")
        print("訓練データ精度：", last_train_acc)
        print("評価データ精度：", last_val_acc)
        print("テストデータ精度：", test_acc)
        print("適合率", prec)        
        print("再現率", recall)
        print("F1", f)
        print("------------------------------------------")
        print("バッチサイズ：",BATCH_SIZE)
        print("学習率：",OPTIMIZER_LEARNING_RATE)
        print("エポック数：",last_epoch_num)
        print("Confusion Matrix",cm)

        #配列宣言
        eval = ['RNN','訓練データ精度','評価データ精度','テストデータ', '適合率','再現率','F1']
        Headdata = ['頭部姿勢あり', last_train_acc, last_val_acc, test_acc, prec, recall, f]
        NoHeaddata = ['頭部姿勢なし',last_train_acc, last_val_acc, test_acc, prec, recall, f]
        
        #配列ループ
        ws.cell(start_row, start_column, "All")
        ws.cell(start_row, start_column+1, i)
        for j in range(0,len(eval)):
            #A列にリストを書き込み
            ws.cell(start_row+j+1,start_column,value = eval[j])
            #B列にリストを書き込み
            if headdatanum == 2:
                ws.cell(start_row+j+1,start_column+1,value = Headdata[j])
            else:
                ws.cell(start_row+j+1,start_column+2, value = NoHeaddata[j])
        ws.cell(start_row+8, start_column, value='エポック数')
        ws.cell(start_row+8, start_column+1, value=last_epoch_num+1)
        wb.save(xlsx_filename)
        start_row += 10
        if i == K-1:#K-分割の最終結果書き込み
            train_mean = statistics.mean(train_acc_list)
            val_mean = statistics.mean(val_acc_list)
            test_mean = statistics.mean(test_acc_list)
            prec_mean = statistics.mean(prec_list)
            recall_mean = statistics.mean(recall_list)
            f_mean = statistics.mean(f_list)
            meanHeaddata = ['頭部姿勢あり', train_mean, val_mean, test_mean, prec_mean, recall_mean, f_mean]
            meanNoHeaddata = ['頭部姿勢なし', train_mean, val_mean, test_mean, prec_mean, recall_mean, f_mean]
            start_row+=2
            ws.cell(start_row, start_column, "All")
            ws.cell(start_row, start_column+1, '平均')
            for i in range(0,len(eval)):
                #A列にリストを書き込み
                ws.cell(start_row+i+1,start_column,value = eval[i])

                #B列にリストを書き込み
                if headdatanum == 2:
                    ws.cell(start_row+i+1,start_column+1,value = meanHeaddata[i])
                else:
                    ws.cell(start_row+i+1,start_column+2, value = meanNoHeaddata[i])
            


column_count=1
row_count=1

TrainandVal(inputfile_name, 2, row_count, column_count)
TrainandVal(inputfile_name, 0, row_count, column_count)
column_count += 4






